#通过 account 的快照，分析结果
"""
账户分析器模块

设计思路:
---------
1. 数据初始化
   - 传入 AccountManager 实例，自动聚合快照为日数据(_daily_assets)
   - 计算交易盈亏记录(_trade_profits)
   - 数据在初始化时固化，后续计算基于此数据

2. 时间区间切片
   - sliced_data: 缓存当前区间的切片数据
   - getTimeRange(period): 设置时间区间，更新缓存
   - _ensure_sliced_data(): 确保切片数据已初始化

3. 使用方式
   - 默认使用全数据: 直接调用指标方法，自动初始化
   - 指定区间: 先调用 getTimeRange('3m')，再调用指标方法
   - 切换区间: 再次调用 getTimeRange() 即可

   示例:
   >>> analyzer = AccountAnalyzer(account=my_account)
   >>> analyzer.volatility()           # 使用全数据
   >>> analyzer.getTimeRange('3m')     # 设置近3月区间
   >>> analyzer.volatility()           # 使用近3月数据
   >>> analyzer.sharpe_ratio()         # 复用缓存的近3月数据

4. 指标分类
   - 资产类指标: 支持区间切片（收益率、波动率、夏普比率等）
   - 交易类指标: 暂不支持区间切片（胜率、盈亏比等）
"""
from collections import defaultdict
from dateutil.relativedelta import relativedelta
import math
import numpy as np
import os
from datetime import datetime, date
from typing import Dict, List, Tuple, Any, Optional, Union
from dataclasses import dataclass
from pathlib import Path
import inspect
from functools import wraps
# [修复] 2026-05-30 _calculate_profit 需要用 OrderSide 枚举做比较
from .account import OrderSide




# ============================================================================
# 指标装饰器
# ============================================================================

def metric(name: str = None, group: str = '', desc: str = '', 
           fmt: str = '.2f', order: int = 99):
    """
    指标装饰器，用于标记分析方法
    
    Args:
        name: 指标名称，默认使用函数名
        group: 指标分组，如 '收益'、'风险'、'交易'，用于自动分组展示
        desc: 指标描述（可选），输入到 notebook metric-desc 标签
        fmt: 输出格式，默认 '.1%'（百分比）。可选 '.2f'、'.1f'
            仅影响 to_notebook / to_excel 的输出，不改变方法返回值
        order: 排序号，用于报告中的顺序
    
    Usage:
        @metric(name='夏普比率', group='风险', fmt='.2f', order=30)
        @metric(name='年化波动率', group='风险', order=20)
    """
    def decorator(func):
        @wraps(func)
        def wrapper(self, *args, **kwargs):
            return func(self, *args, **kwargs)
        
        wrapper._is_metric = True
        wrapper._metric_name = name or func.__name__
        wrapper._metric_group = group
        wrapper._metric_desc = desc
        wrapper._metric_fmt = fmt
        wrapper._metric_order = order
        return wrapper
    return decorator



# ============================================================================
# 账户分析类
# ============================================================================

class AccountAnalyzer:
    """
    账户分析器 — 基于快照的绩效分析、指标计算与报告输出

    核心原则：计算层返回纯数字，呈现层通过 @metric.fmt 控制格式。
    新增指标只需加 @metric(group='收益/风险/交易')，to_notebook/to_excel 自动拾取。

    AccountAnalyzer
    ├── __init__()                      # 三种数据源: account / daily_assets(dict) / daily_assets(list)
    │
    ├── [属性]                          # 对外只读 copy
    │   ├── daily_assets                #  Dict[date, float] 每日净值
    │   └── trade_profits               #  List[Dict] 逐笔盈亏 (symbol/profit/open_time/...)
    │
    ├── [基准对比]                      # 2026-06-02 新增
    │   └── set_benchmark()             #  注入基准日净值 → to_notebook() 懒计算对比 Table + 净值叠加 + 超额曲线
    │
    ├── [时间区间切片]                  # 缓存 sliced_data，各指标共享预计算结果
    │   ├── getTimeRange()              #  设置区间: 默认/'all'/'3m'/'1y'/自定义起止 → 预计算 values/returns
    │   └── _ensure_sliced_data()       #  惰性初始化，首次调用指标时自动 fallback 到全区间
    │
    ├── [指标计算 @metric]              # 声明式驱动，按 group 分组 (收益/风险/交易)
    │   ├── [收益] return_rate          #  累计收益率
    │   ├── [收益] annualized_return    #  年化收益率
    │   ├── [风险] volatility           #  年化波动率
    │   ├── [风险] max_drawdown         #  最大回撤 (返回 Tuple[rate, peak_date, trough_date])
    │   ├── [风险] sharpe_ratio         #  夏普比率
    │   ├── [风险] sortino_ratio        #  索提诺比率 (仅下行风险)
    │   ├── [风险] upi                  #  溃疡绩效指数 (Ulcer Performance Index)
    │   ├── [风险] ulcer_index          #  溃疡指数
    │   ├── [风险] var                  #  VaR(95%)
    │   ├── [风险] cvar                 #  CVaR(95%)
    │   ├── [交易] win_rate             #  胜率
    │   ├── [交易] avg_profit_loss_ratio #  平均盈亏比
    │   ├── [交易] avg_holding_period   #  平均持仓天数
    │   ├── [交易] kelly_criterion      #  凯利最优仓位
    │   ├── [交易] kelly_fraction       #  半凯利仓位
    │   ├── avg_profit()                #  (辅助) 平均盈利, 支持 amount/percentage 模式
    │   └── avg_loss()                  #  (辅助) 平均亏损
    │
    ├── [指标收集]                      # _collect_metrics() 遍历 @metric → to_notebook/to_excel 自动拾取
    │   ├── _collect_metrics()          #  收集所有 @metric 方法 → {name: {value, group, fmt, desc}}
    │   ├── metrics()                   #  公开入口
    │   └── returns()                   #  批量多区间收益率 ('1m,3m,1y')
    │
    ├── [查询]
    │   ├── get_daily_total_assets()    #  返回内部 _daily_assets
    │   ├── get_largest_profit_trades() #  盈利 Top N
    │   └── get_largest_loss_trades()   #  亏损 Top N
    │
    ├── [导出]
    │   ├── to_notebook()               #  HTML 交互式报告 (无基准→三段式 / 有基准→对比前置)
    │   └── to_excel()                  #  Excel (Sheet: 回测指标/每日资产/交易记录)
    │
    └── [底层]
        ├── _aggregate_daily_assets()   #  快照列表 → Dict[date, nav] (同日期取最后)
        ├── _calculate_profit()         #  FIFO 匹配计算逐笔盈亏
        └── _get_caller_dir()           #  调用者目录 (报告输出路径基准)
    """

    def __init__(self, account=None, daily_assets=None):
        """
        初始化账户分析器
        
        支持两种数据输入方式：
        1. 传入 account 实例：自动聚合快照为日数据，并计算交易盈亏
        2. 传入 daily_assets：直接使用外部提供的日数据（不计算交易盈亏）
        
        Args:
            account: AccountManager 实例，提供原始快照数据
                   如果提供，将自动调用 _aggregate_daily_assets 聚合日数据
            daily_assets: 外部每日资产数据，支持两种格式：
                         - List[Dict]: [{'date': date(2024,1,1), 'assets': 100000}, ...]
                         - Dict[date, float]: {date(2024,1,1): 100000, ...}
                        如果提供，将优先使用此数据而非 account 的快照
            
        Example:
            >>> # 方式 1：使用 account 实例
            >>> analyzer = AccountAnalyzer(account=my_account)
            >>> 
            >>> # 方式 2：使用外部日数据（List 格式）
            >>> daily_data = [
            ...     {'date': date(2024, 1, 1), 'assets': 100000},
            ...     {'date': date(2024, 1, 2), 'assets': 101000},
            ... ]
            >>> analyzer = AccountAnalyzer(daily_assets=daily_data)
            >>> 
            >>> # 方式 3：使用外部日数据（Dict 格式）
            >>> daily_dict = {date(2024,1,1): 100000, date(2024,1,2): 101000}
            >>> analyzer = AccountAnalyzer(daily_assets=daily_dict)
        """
        self.account = account
        self.risk_free_rate = 0.02
        self.sliced_data = None
        # [新增] 2026-06-02 基准对比支持
        #   通过 set_benchmark() 注入基准日净值数据，to_notebook() 自动生成对比 section
        #   对比指标通过懒计算与策略日期对齐，避免时间切片不一致
        self._bench_assets = None   # Dict[date, float] 基准每日净值
        self._bench_label = ""      # str 基准名称
        
        if daily_assets is not None:
            if isinstance(daily_assets, dict):
                self._daily_assets = daily_assets
            else:
                self._daily_assets = {item['date']: item['assets'] for item in daily_assets}
            self._trade_profits = []
        elif account:
            self._daily_assets = self._aggregate_daily_assets(account.snapshots)
            self._trade_profits = self._calculate_profit(account.trade_records)
        else:
            self._daily_assets = {}
            self._trade_profits = []

        self.base_dir = self._get_caller_dir()

    @property
    def daily_assets(self) -> Dict:
        """
        获取每日资产净值的副本
        
        返回内部存储的每日资产数据的浅拷贝，防止外部修改
        数据格式：{date: nav}
        
        Returns:
            Dict[date, float]: 每日资产净值字典的副本
            
        Example:
            >>> assets = analyzer.daily_assets
            >>> print(f"最新资产：{assets[max(assets.keys())]}")
        """
        return self._daily_assets.copy()

    @property
    def trade_profits(self) -> List:
        """
        获取交易盈亏记录的副本
        
        返回内部存储的交易盈亏数据的浅拷贝，防止外部修改
        每笔交易记录包含：symbol, profit, open_time, close_time 等信息
        
        Returns:
            List[Dict]: 交易盈亏记录列表的副本
            
        Example:
            >>> profits = analyzer.trade_profits
            >>> total_profit = sum(t['profit'] for t in profits)
            >>> print(f"总盈利：{total_profit}")
        """
        return self._trade_profits.copy()

    # ------------------------------------------------------------------------
    # 基准对比支持
    # ------------------------------------------------------------------------
    # [新增] 2026-06-02 注入基准日净值数据，to_notebook() 自动生成对比 section
    #   数据格式与 _daily_assets 保持一致：Dict[date, float] 或 List[{'date': d, 'assets': v}]
    #   指标计算采用懒加载：在 to_notebook() 里按策略日期对齐后，创建临时 AccountAnalyzer
    #   跑一次 metrics()，保证基准指标与策略指标完全同构（同一套 @metric 方法）
    #   方法返回 self 以支持链式调用

    def set_benchmark(self, daily_assets, label: str = "基准") -> 'AccountAnalyzer':
        """
        注入基准日净值数据，开启对比分析模式。

        基准指标在 to_notebook() 中懒计算：
        - 按策略当前日期范围对齐
        - 内部运行 AccountAnalyzer(daily_assets=aligned) 获取同构指标
        - 自动生成对比 Table + 净值叠加图 + 超额曲线

        Args:
            daily_assets: 基准每日资产净值，支持两种格式：
                         - Dict[date, float]: {date(2020,1,2): 100000, ...}
                         - List[Dict]: [{'date': date(2020,1,2), 'assets': 100000}, ...]
            label: 基准名称，显示在图例和指标中

        Returns:
            self，支持链式调用

        Example:
            bench_analyzer = run_backtest(bars, BenchHolder(), '基准')
            strategy_analyzer.set_benchmark(bench_analyzer.daily_assets, '国证A指')
            strategy_analyzer.to_notebook("策略 vs 基准")
        """
        if isinstance(daily_assets, dict):
            self._bench_assets = daily_assets
        else:
            self._bench_assets = {item['date']: item['assets'] for item in daily_assets}
        self._bench_label = label
        return self

    # ------------------------------------------------------------------------
    # 链式调用方法（支持 @metric 装饰器）
    # ------------------------------------------------------------------------

    _periods_map = {
        '1m': relativedelta(months=1),
        '3m': relativedelta(months=3),
        '6m': relativedelta(months=6),
        '1y': relativedelta(years=1),
        '2y': relativedelta(years=2),
        '3y': relativedelta(years=3),
        '5y': relativedelta(years=5),
        'all': None
    }

    def _ensure_sliced_data(self) -> Optional[Dict]:
        """
        确保切片数据已初始化
        
        如果 sliced_data 为 None，自动调用 getTimeRange() 初始化全数据
        
        Returns:
            Dict: 切片数据，格式为：
                {
                    'startDate': date,
                    'endDate': date,
                    'baseDate': date,
                    'daily_assets': Dict
                }
                数据不足时返回 None
        """
        if self.sliced_data is None:
            self.getTimeRange()
        return self.sliced_data

    def getTimeRange(self, period_or_start=None, end=None) -> Optional[Dict]:
        """
        获取时间区间信息并缓存切片数据
        
        Args:
            period_or_start: 
                - None: 使用全部数据
                - str: 周期字符串，如 '1m', '3m', '6m', '1y', 'all'
                - date: 自定义开始日期
            end: 结束日期（当 period_or_start 为 date 时使用）
        
        Returns:
            Dict: 时间区间信息，格式为：
                {
                    'startDate': date,      # 区间起始日期
                    'endDate': date,        # 区间结束日期
                    'baseDate': date,       # 基准日日期
                    'daily_assets': Dict    # 切片后的日资产数据
                }
                数据不足时返回 None
            
        Example:
            >>> timeRange = analyzer.getTimeRange('3m')
            >>> print(timeRange['startDate'], timeRange['endDate'])
            >>> timeRange = analyzer.getTimeRange(date(2024,1,1), date(2024,6,30))
        """
        if not self._daily_assets:
            return None
        
        dates = sorted(self._daily_assets.keys())
        if len(dates) < 2:
            return None
        
        all_benchmark = dates[0]
        all_start = dates[1]
        all_end = dates[-1]
        
        if period_or_start is None:
            start_date = all_start
            end_date = all_end
            benchmark_date = all_benchmark
        elif isinstance(period_or_start, str):
            if period_or_start == 'all':
                start_date = all_start
                end_date = all_end
                benchmark_date = all_benchmark
            else:
                delta = self._periods_map.get(period_or_start)
                if delta is None:
                    start_date = all_start
                    end_date = all_end
                    benchmark_date = all_benchmark
                else:
                    calculated_start = dates[-1] - delta
                    valid_dates = [d for d in dates if d <= calculated_start]
                    raw_start = max(valid_dates) if valid_dates else all_start
                    
                    if raw_start == all_start:
                        start_date = all_start
                        end_date = all_end
                        benchmark_date = all_benchmark
                    else:
                        prev_dates = [d for d in dates if d < raw_start]
                        benchmark_date = max(prev_dates) if prev_dates else all_benchmark
                        start_date = raw_start
                        end_date = all_end
        else:
            raw_start = period_or_start or all_start
            raw_end = end or all_end
            
            if raw_start < all_start:
                raw_start = all_start
            
            prev_dates = [d for d in dates if d < raw_start]
            if prev_dates:
                benchmark_date = max(prev_dates)
            else:
                benchmark_date = all_benchmark
            
            start_date = raw_start
            end_date = raw_end
        
        sliced_assets = {
            d: v for d, v in self._daily_assets.items()
            if benchmark_date <= d <= end_date
        }
        
        # ====================================================================
        # 性能优化：预计算基础数据
        # ====================================================================
        # 在切片数据时一次性计算所有指标需要的基础数据，避免重复计算
        # 
        # 优化效果：
        # 1. 日收益率只需计算一次，volatility/var/cvar/sortino_ratio 等方法共享
        # 2. 使用 NumPy 向量化计算，比循环快 10-50 倍
        # 3. 后续指标方法直接使用预计算结果，无需重复排序和计算
        # ====================================================================
        
        # 排序日期列表（用于后续索引）
        dates = sorted(sliced_assets.keys())
        
        # 资产值数组（NumPy 向量化计算的基础）
        values = np.array([sliced_assets[d] for d in dates])
        
        # 日收益率数组（向量化计算：今日/昨日 - 1）
        # 公式：(values[1:] - values[:-1]) / values[:-1]
        # 等价于：[values[i]/values[i-1] - 1 for i in range(1, len(values))]
        # 但 NumPy 向量化版本快 20-50 倍
        daily_returns = (values[1:] - values[:-1]) / values[:-1] if len(values) > 1 else np.array([])
        
        self.sliced_data = {
            'startDate': start_date,
            'endDate': end_date,
            'baseDate': benchmark_date,
            'daily_assets': sliced_assets,  # 原始数据（兼容旧代码）
            'dates': dates,                  # 排序后的日期列表
            'values': values,                # 资产值数组（用于最大回撤、Ulcer Index 等）
            'returns': daily_returns         # 日收益率数组（用于波动率、VaR、CVaR、索提诺比率等）
        }
        
        return self.sliced_data

    @metric(name='累计收益率', group='收益', fmt='.1%', desc='统计期间内的总收益率', order=10)
    def return_rate(self) -> Optional[float]:
        """计算累计收益率（链式调用版本）"""
        if not self._daily_assets:
            return None
        
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        sliced_data = sliced_data_info['daily_assets']
        
        if len(sliced_data) < 2:
            return None
        
        dates = sorted(sliced_data.keys())
        benchmark_value = sliced_data[dates[0]]
        end_value = sliced_data[dates[-1]]
        
        if benchmark_value == 0:
            return None
        
        return (end_value - benchmark_value) / benchmark_value

    @metric(name='年化收益率', group='收益', fmt='.1%', desc='将收益率换算为年化基准', order=11)
    def annualized_return(self) -> Optional[float]:
        """计算年化收益率"""
        interval_return = self.return_rate()
        if interval_return is None:
            return None
        
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        sliced_data = sliced_data_info['daily_assets']
        
        dates = sorted(sliced_data.keys())
        trading_days = len(dates) - 1
        
        if trading_days <= 0:
            return 0
        if interval_return <= -1:
            return None
        
        return ((1 + interval_return) ** (252 / trading_days)) - 1

    @metric(name='年化波动率', group='风险', fmt='.1%', desc='衡量资产价格的波动程度', order=22)
    def volatility(self) -> Optional[float]:
        """计算年化波动率"""
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        returns = sliced_data_info.get('returns')
        if returns is None or len(returns) < 2:
            return None
        
        return np.std(returns) * np.sqrt(252)

    @metric(name='夏普比率', group='风险', fmt='.2f', desc='每承担一单位风险获得的超额收益', order=21)
    def sharpe_ratio(self) -> Optional[float]:
        """计算夏普比率"""
        annualized = self.annualized_return()
        vol = self.volatility()
        
        if annualized is None or vol is None or vol == 0:
            return None
        
        return (annualized - self.risk_free_rate) / vol

    @metric(name='最大回撤', group='风险', fmt='.1%', desc='历史最大亏损幅度', order=20)
    def max_drawdown(self) -> Optional[Tuple[float, date, date]]:
        """计算最大回撤"""
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        values = sliced_data_info.get('values')
        dates = sliced_data_info.get('dates')
        
        if values is None or len(values) < 2:
            return None
        
        cumulative = values / values[0]
        running_max = np.maximum.accumulate(cumulative)
        drawdown = (running_max - cumulative) / running_max
        
        max_dd_idx = np.argmax(drawdown)
        max_dd = drawdown[max_dd_idx]
        
        if max_dd == 0:
            return None
        
        peak_idx = np.argmax(cumulative[:max_dd_idx + 1])
        
        return max_dd, dates[peak_idx], dates[max_dd_idx]

    @metric(name='VaR(95%) / 风险价值', group='风险', fmt='.1%', desc='95% 置信度下的最大可能损失', order=24)
    def var(self, confidence: float = 0.95) -> Optional[float]:
        """计算风险价值"""
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        returns = sliced_data_info.get('returns')
        if returns is None or len(returns) < 2:
            return None
        
        return -np.percentile(returns, (1 - confidence) * 100)

    @metric(name='CVaR(95%) / 条件风险价值', group='风险', fmt='.1%', desc='超过 VaR 阈值的平均损失', order=25)
    def cvar(self, confidence: float = 0.95) -> Optional[float]:
        """计算条件风险价值"""
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        returns = sliced_data_info.get('returns')
        if returns is None or len(returns) < 2:
            return None
        
        index = int((1 - confidence) * len(returns))
        if index < 1:
            index = 1
        
        sorted_returns = np.sort(returns)
        return -np.mean(sorted_returns[:index])

    @metric(name='Ulcer Index / 溃疡指数', group='风险', fmt='.2f', desc='衡量回撤深度和持续时间的综合指标', order=27)
    def ulcer_index(self) -> Optional[float]:
        """计算溃疡指数"""
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        values = sliced_data_info.get('values')
        
        if values is None or len(values) < 2:
            return None
        
        cumulative = values / values[0]
        running_max = np.maximum.accumulate(cumulative)
        drawdown_pct = ((running_max - cumulative) / running_max) * 100
        
        return np.sqrt(np.mean(drawdown_pct ** 2))

    @metric(name='索提诺比率', group='风险', fmt='.2f', desc='只考虑下行风险的夏普比率改进版', order=23)
    def sortino_ratio(self, risk_free_rate: float = 0.02) -> Optional[float]:
        """计算索提诺比率"""
        annualized_return = self.annualized_return()
        if annualized_return is None:
            return None
        
        sliced_data_info = self._ensure_sliced_data()
        if sliced_data_info is None:
            return None
        
        returns = sliced_data_info.get('returns')
        if returns is None or len(returns) < 2:
            return None
        
        negative_returns = returns[returns < 0]
        if len(negative_returns) == 0:
            return float('inf')
        
        downside_variance = np.mean(returns ** 2 * (returns < 0))
        downside_deviation = np.sqrt(downside_variance)
        annualized_downside_deviation = downside_deviation * np.sqrt(252)
        
        if annualized_downside_deviation == 0:
            return float('inf')
        
        return (annualized_return - risk_free_rate) / annualized_downside_deviation

    @metric(name='UPI / 溃疡绩效指数', group='风险', fmt='.2f', desc='用溃疡指数调整的风险收益比', order=26)
    def upi(self, risk_free_rate: float = 0.02) -> Optional[float]:
        """计算 Ulcer Performance Index"""
        annualized_return = self.annualized_return()
        ulcer_idx = self.ulcer_index()
        
        if annualized_return is None or ulcer_idx is None or ulcer_idx == 0:
            return None
        
        return (annualized_return - risk_free_rate) / (ulcer_idx / 100)

    @metric(name='胜率', group='交易', fmt='.1%', desc='盈利交易次数占总交易次数的比例', order=40)
    def win_rate(self) -> Optional[float]:
        """计算胜率"""
        if not self._trade_profits:
            return None
        
        wins = sum(1 for t in self._trade_profits if t['profit'] > 0)
        return wins / len(self._trade_profits)

    def avg_profit(self, mode: str = 'amount') -> Optional[float]:
        """
        计算平均盈利
        
        Args:
            mode: 计算模式
                  - 'amount': 金额模式，计算平均盈利金额
                  - 'percentage': 百分比模式，计算平均盈利占本金的比例
            
        Returns:
            float: 平均盈利
                  无盈利交易时返回 None
        """
        profitable_trades = [t for t in self._trade_profits if t['profit'] > 0]
        if not profitable_trades:
            return None
        
        if mode == 'amount':
            profits = [t['profit'] for t in profitable_trades]
        elif mode == 'percentage':
            profits = [
                t['profit'] / (abs(t['volume']) * t['open_price'])
                for t in profitable_trades
            ]
        else:
            raise ValueError("mode 必须是 'amount' 或 'percentage'")
        
        return sum(profits) / len(profits)

    def avg_loss(self, mode: str = 'amount') -> Optional[float]:
        """
        计算平均亏损
        
        Args:
            mode: 计算模式
                  - 'amount': 金额模式，计算平均亏损金额
                  - 'percentage': 百分比模式，计算平均亏损占本金的比例
            
        Returns:
            float: 平均亏损（负数）
                  无亏损交易时返回 None
        """
        loss_trades = [t for t in self._trade_profits if t['profit'] < 0]
        if not loss_trades:
            return None
        
        if mode == 'amount':
            losses = [t['profit'] for t in loss_trades]
        elif mode == 'percentage':
            losses = [
                t['profit'] / (abs(t['volume']) * t['open_price'])
                for t in loss_trades
            ]
        else:
            raise ValueError("mode 必须是 'amount' 或 'percentage'")
        
        return sum(losses) / len(losses)

    @metric(name='平均盈亏比', group='交易', fmt='.2f', desc='平均盈利与平均亏损的比值', order=41)
    def avg_profit_loss_ratio(self) -> Optional[float]:
        """计算平均盈亏比"""
        avg_profit = self.avg_profit()
        avg_loss = self.avg_loss()
        
        if avg_profit is None or avg_loss is None or avg_loss == 0:
            return None
        
        return abs(avg_profit / avg_loss)

    @metric(name='平均持仓天数', group='交易', fmt='.1f', desc='所有交易的平均持仓天数', order=42)
    def avg_holding_period(self) -> Optional[float]:
        """计算平均持仓时间"""
        if not self._trade_profits:
            return None
        
        total_days = sum((t['close_time'] - t['open_time']).days for t in self._trade_profits)
        return total_days / len(self._trade_profits)

    @metric(name='凯利公式最优仓位', group='交易', fmt='.1%', desc='根据胜率和盈亏比计算的最优仓位比例', order=50)
    def kelly_criterion(self) -> Optional[float]:
        """计算凯利公式最优仓位"""
        if not self._trade_profits:
            return None
        
        win_rate_val = self.win_rate()
        if win_rate_val is None:
            return None
        
        avg_profit = self.avg_profit()
        avg_loss = self.avg_loss()
        
        if avg_profit is None or avg_loss is None or avg_loss == 0:
            return None
        
        profit_loss_ratio = abs(avg_profit / avg_loss)
        return win_rate_val - (1 - win_rate_val) / profit_loss_ratio

    @metric(name='半凯利仓位', group='交易', fmt='.1%', desc='凯利公式最优仓位的50%', order=51)
    def kelly_fraction(self, fraction: float = 0.5) -> Optional[float]:
        """计算凯利分数仓位"""
        kelly = self.kelly_criterion()
        if kelly is None:
            return None
        return kelly * fraction

    def _collect_metrics(self) -> Dict:
        """收集所有 @metric 装饰器标记的方法"""
        metrics = {}
        for name in dir(self):
            if name.startswith('_'):
                continue
            attr = getattr(self, name, None)
            if callable(attr) and hasattr(attr, '_is_metric'):
                try:
                    value = attr()
                    metrics[attr._metric_name] = {
                        'name': attr._metric_name,
                        'value': value,
                        'group': getattr(attr, '_metric_group', ''),
                        'desc': attr._metric_desc,
                        'fmt': getattr(attr, '_metric_fmt', '.1%'),
                        'order': attr._metric_order,
                        'method': name
                    }
                except:
                    pass
        return metrics

    def metrics(self) -> Dict:
        """获取当前区间的所有指标"""
        return self._collect_metrics()

    def returns(self, periods: str = None) -> Dict:
        """批量计算多区间收益率"""
        if periods is None:
            return {'default': self.return_rate()}
        
        period_list = [p.strip() for p in periods.split(',')]
        result = {}
        
        for p in period_list:
            timeRange = self.getTimeRange(p)
            if timeRange:
                sliced_data = timeRange['daily_assets']
                if len(sliced_data) >= 2:
                    dates = sorted(sliced_data.keys())
                    benchmark_value = sliced_data[dates[0]]
                    end_value = sliced_data[dates[-1]]
                    if benchmark_value != 0:
                        result[p] = (end_value - benchmark_value) / benchmark_value
                    else:
                        result[p] = None
                else:
                    result[p] = None
            else:
                result[p] = None
        
        return result

    # ------------------------------------------------------------------------
    # 查询方法
    # ------------------------------------------------------------------------

    def get_daily_total_assets(self) -> Dict:
        """
        获取每日资产净值
        
        返回聚合后的每日资产数据，可用于绘制净值曲线
        
        Returns:
            Dict[date, float]: {日期：资产净值} 字典
                              日期为 date 对象，净值为 float
            
        Example:
            >>> assets = analyzer.get_daily_total_assets()
            >>> for date, value in sorted(assets.items()):
            ...     print(f"{date}: {value}")
        """
        return self._daily_assets

    def get_largest_profit_trades(self, n: int) -> List:
        """
        获取盈利最大的 N 笔交易
        
        按盈利金额降序排序，返回前 N 笔交易记录
        
        Args:
            n: 返回的交易数量
            
        Returns:
            List[Dict]: 交易记录列表，每笔交易包含：
                       - symbol: 标的代码
                       - profit: 盈亏金额
                       - open_time, close_time: 开平仓时间
                       - open_price, close_price: 开平仓价格
                       - volume: 交易数量
                       按盈利降序排序
            
        Example:
            >>> top5 = analyzer.get_largest_profit_trades(5)
            >>> for trade in top5:
            ...     print(f"{trade['symbol']}: 盈利 {trade['profit']}")
        """
        if not self._trade_profits or n <= 0:
            return []
        return sorted(self._trade_profits, key=lambda t: t['profit'], reverse=True)[:n]

    def get_largest_loss_trades(self, n: int) -> List:
        """
        获取亏损最大的 N 笔交易
        
        按亏损金额升序排序（亏损最多的在前），返回前 N 笔交易记录
        用于分析策略的失败案例，找出需要改进的地方
        
        Args:
            n: 返回的交易数量
            
        Returns:
            List[Dict]: 交易记录列表，按亏损升序排序
                       （亏损最多的交易在最前面）
            
        Example:
            >>> worst5 = analyzer.get_largest_loss_trades(5)
            >>> for trade in worst5:
            ...     print(f"{trade['symbol']}: 亏损 {abs(trade['profit'])}")
        """
        if not self._trade_profits or n <= 0:
            return []
        return sorted(self._trade_profits, key=lambda t: t['profit'])[:n]

    # ------------------------------------------------------------------------
    # 统一数据准备层
    # ------------------------------------------------------------------------
    @staticmethod
    def _fmt(val, t='.1%'):
        """格式化指标值，Python format 规范 ('.1%', '.2%', '.2f')"""
        if val is None: return 'N/A'
        if isinstance(val, tuple): val = val[0]
        if t.endswith('%'):
            d = int(t[1:-1])
            return f"{val*100:.{d}f}%"
        if t.endswith('f'):
            d = int(t[1:-1])
            return f"{val:.{d}f}"
        return str(val)

    def _build_report_data(self):
        """[重构] 2026-06-02 统一数据准备，消除 to_notebook/to_excel 重复逻辑
        
        Returns:
            dict: {
                'nav_values': list, 'dates': list, 'info_m': dict,
                'grouped': dict, 'base_items': list, 'has_records': bool,
                'has_bench': bool, 'cmp_rows': list, 'bench_label': str,
                'common_dates': list, 'strat_vals': np.array, 'bench_vals': np.array,
            }
        """
        dates = sorted(self._daily_assets.keys())
        assets = self._daily_assets
        nav_values = [assets[d] for d in dates] if len(dates) >= 2 else []
        all_metrics = self.metrics()

        # 基础信息
        info_m = {}
        if dates and len(dates) >= 2:
            info_m['开始日期'] = dates[1].strftime('%Y-%m-%d')
            info_m['结束日期'] = dates[-1].strftime('%Y-%m-%d')
        initial_cash = self.account.snapshots[0].cash if self.account and self.account.snapshots else 0
        final_nav = self.account.snapshots[-1].nav if self.account and self.account.snapshots else 0
        if initial_cash > 0:
            info_m['初始资金'] = f"{initial_cash:,.0f}"
        if final_nav > 0:
            info_m['最终资产'] = f"{final_nav:,.0f}"

        # 指标分组
        has_records = bool(self.account and self.account.trade_records)
        grouped = {}
        for m in all_metrics.values():
            g = m.get('group', '')
            if not g:
                continue
            val = m['value']
            if isinstance(val, tuple):
                val = val[0]
            if val is not None and not (isinstance(val, float) and val == float('inf')):
                grouped.setdefault(g, []).append(
                    (m['order'], m['name'], self._fmt(val, m.get('fmt', '.1%')), m.get('desc', ''))
                )
        for items in grouped.values():
            items.sort(key=lambda x: x[0])

        # 基础指标（无基准时使用）
        base_items = []
        for k, v in info_m.items():
            base_items.append({'name': k, 'value': v})
        for group_name in ('收益', '风险'):
            if group_name in grouped:
                for _, name, val, desc in grouped[group_name]:
                    item = {'name': name, 'value': val}
                    if desc:
                        item['desc'] = desc
                    base_items.append(item)

        # 交易指标：盈亏比附加百分比
        if has_records and '交易' in grouped:
            avg_p = self.avg_profit(mode='percentage')
            avg_l = self.avg_loss(mode='percentage')
            if avg_p is not None and avg_l is not None:
                for items in grouped['交易']:
                    if items[1] == '平均盈亏比':
                        idx = grouped['交易'].index(items)
                        grouped['交易'][idx] = (items[0], items[1],
                            f"{items[2]}（盈 +{avg_p*100:.1f}% / 亏 {avg_l*100:.1f}%）", items[3])
                        break

        # 基准对比
        cmp_rows = []
        has_bench_data = False
        common_dates = []
        strat_vals = np.array([])
        bench_vals = np.array([])
        bench_label = self._bench_label or '基准'
        if self._bench_assets:
            bench_all = dict(self._bench_assets)
            strat_dates = sorted(self._daily_assets.keys())
            bench_dates = sorted(bench_all.keys())
            for d in strat_dates:
                if d < bench_dates[0]:
                    bench_all[d] = bench_all[bench_dates[0]]
            common_dates = sorted(set(self._daily_assets.keys()) & set(bench_all.keys()))
            if len(common_dates) >= 2:
                has_bench_data = True
                strat_vals = np.array([self._daily_assets[d] for d in common_dates])
                bench_vals = np.array([bench_all[d] for d in common_dates])
                bench_aligned = {d: bench_all[d] for d in common_dates}
                bench_temp = AccountAnalyzer(daily_assets=bench_aligned)
                bench_all_m = bench_temp.metrics()

                sorted_metrics = sorted(all_metrics.values(), key=lambda m: m.get('order', 99))
                for s_m in sorted_metrics:
                    name = s_m['name']
                    if name not in bench_all_m:
                        continue
                    b_m = bench_all_m[name]
                    if b_m.get('group') == '交易':
                        continue
                    s_v = s_m['value']
                    b_v = b_m['value']
                    if isinstance(s_v, tuple): s_v = s_v[0]
                    if isinstance(b_v, tuple): b_v = b_v[0]
                    fmt = s_m.get('fmt', '.1%')
                    cmp_rows.append({
                        '指标': s_m['name'],
                        '策略': self._fmt(s_v, fmt),
                        bench_label: self._fmt(b_v, fmt),
                    })
                # 超额指标
                s_total = strat_vals[-1] / strat_vals[0] - 1
                b_total = bench_vals[-1] / bench_vals[0] - 1
                excess_total = (1 + s_total) / (1 + b_total) - 1 if b_total > -1 else None
                days_n = len(common_dates) - 1
                if excess_total is not None:
                    cmp_rows.append({'指标': '超额收益', '策略': self._fmt(excess_total, '.1%'), bench_label: '—'})
                    if days_n > 0:
                        ann_excess = (1 + excess_total) ** (252 / days_n) - 1
                        cmp_rows.append({'指标': '年化超额', '策略': self._fmt(ann_excess, '.1%'), bench_label: '—'})
                        s_rets = (strat_vals[1:] - strat_vals[:-1]) / strat_vals[:-1]
                        b_rets = (bench_vals[1:] - bench_vals[:-1]) / bench_vals[:-1]
                        excess_daily = s_rets - b_rets
                        te = np.std(excess_daily) * np.sqrt(252)
                        ir = ann_excess / te if te > 0 else None
                        if ir is not None:
                            cmp_rows.append({'指标': '信息比率', '策略': f'{ir:.2f}', bench_label: '—'})
                        cmp_rows.append({'指标': '跟踪误差', '策略': self._fmt(te, '.1%'), bench_label: '—'})
                        day_win = np.mean(excess_daily > 0)
                        cmp_rows.append({'指标': '日超额胜率', '策略': self._fmt(day_win, '.1%'), bench_label: '—'})

        has_bench = bool(self._bench_assets and has_bench_data)

        return {
            'nav_values': nav_values, 'dates': dates, 'info_m': info_m,
            'grouped': grouped, 'base_items': base_items, 'has_records': has_records,
            'has_bench': has_bench, 'cmp_rows': cmp_rows, 'bench_label': bench_label,
            'common_dates': common_dates, 'strat_vals': strat_vals, 'bench_vals': bench_vals,
        }

    # ------------------------------------------------------------------------
    # 导出方法
    # ------------------------------------------------------------------------

    def to_notebook(self, title: str = "回测报告"):
        """导出为 Notebook 交互式报告，自动保存 HTML 文件

        遵循 notebook 推荐的 section 层次结构：
        - 顶层 KPI 卡片（无 section 包裹，一眼掌握全局）
        - 基础信息 / 指标汇总 / 收益分析(图表) / 交易明细 四大章节

        Args:
            title: 报告标题

        Returns:
            str: 输出的 HTML 文件路径

        Example:
            analyzer.to_notebook("动量策略")
        """
        # [重构] 2026-05-30 按 notebook 推荐层次结构重组：
        #   顶层KPI → 基础信息(日期/资金) → 指标汇总(收益+风险+交易) →
        #   收益分析(净值+回撤图表) → 交易明细(表格,折叠)
        from notebook import Notebook

        nb = Notebook(title)
        nb.base_dir = self.base_dir

        # 统一数据准备
        d = self._build_report_data()

        # ═══════════════════════════
        # 模块1：指标分析
        # ═══════════════════════════
        with nb.section("指标分析"):
            if d['has_bench']:
                nb.table(d['cmp_rows'], title=f"策略 vs {d['bench_label']} 指标对比",
                         columns=['指标', '策略', d['bench_label']])
            else:
                nb.metrics(d['base_items'], columns=5)

            if d['has_records'] and '交易' in d['grouped']:
                trade_m = [{'指标': n, '数值': v} for _, n, v, _ in d['grouped']['交易']]
                nb.table(trade_m, columns=['指标', '数值'])

        # ═══════════════════════════
        # 模块2：收益走势图
        # ═══════════════════════════
        with nb.section("收益走势图"):
            if d['has_bench']:
                nb.chart('perf', {
                    'xAxis': [dt.strftime('%Y-%m-%d') for dt in d['common_dates']],
                    'series': [
                        {'name': '策略', 'data': d['strat_vals'].tolist()},
                        {'name': d['bench_label'], 'data': d['bench_vals'].tolist()},
                    ],
                }, height='500px',
                    series_opts={'is_smooth': True},
                    datazoom_opts=[{'type_': 'slider', 'range_start': 0, 'range_end': 100}])
            elif d['nav_values']:
                nb.chart('perf', {
                    'xAxis': [dt.strftime('%Y-%m-%d') for dt in d['dates']],
                    'series': [{'name': '策略', 'data': d['nav_values']}],
                }, height='500px',
                    series_opts={'is_smooth': True},
                    datazoom_opts=[{'type_': 'slider', 'range_start': 0, 'range_end': 100}])

        # ═══════════════════════════════════════════
        # 模块3：交易记录
        # ═══════════════════════════════════════════
        if d['has_records']:
            trades = []
            has_notes = any(getattr(t, 'note', '') for t in self.account.trade_records)
            for t in self.account.trade_records:
                row = {
                    '日期': t.created_at.strftime('%Y-%m-%d %H:%M'),
                    '标的': t.symbol,
                    '方向': '买入' if t.side == 1 else '卖出',
                    '价格': t.price,
                    '数量': t.volume,
                    '金额': round(t.amount, 2),
                    '手续费': round(t.fee, 2),
                }
                if has_notes:
                    row['备注'] = getattr(t, 'note', '')
                trades.append(row)
            with nb.section("交易记录"):
                nb.table(trades, page={'size': 10})

        return nb.export_html()

    def to_excel(self, report_name: str = "回测报告", output_dir: str = "."):
        """导出 Excel 文件，与 notebook 报告结构对应

        Sheet 结构:
            回测指标 — 按 group（基础/收益/风险/交易）分组展示
            每日资产 — 日期 + 现金 + 持仓市值 + 总净值
            交易记录 — 全部成交明细（日期/标的/方向/价格/数量/金额/手续费/备注）

        Args:
            report_name: 报告名称，用于生成文件名
                       格式：{report_name}_{YYYYMMDD_HHMM}.xlsx
            output_dir: 输出目录（相对路径，基于调用者所在目录）
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # 统一数据准备
        d = self._build_report_data()

        # 转换为 Excel 格式（key: notebook 格式 → Excel 格式）
        base_rows = [{'指标名称': it['name'], '数值': it['value']} for it in d['base_items']]
        groups = {}
        for g_name, items in d['grouped'].items():
            groups[g_name] = [{'指标名称': n, '数值': v, '说明': desc, 'order': o}
                              for o, n, v, desc in items]

        # 基准对比转为 Excel 格式
        cmp_rows = []
        bench_label = d['bench_label']
        has_bench = d['has_bench']
        has_records = d['has_records']
        if has_bench:
            for row in d['cmp_rows']:
                excel_row = {'指标名称': row['指标'], '策略': row['策略']}
                excel_row[d['bench_label']] = row[d['bench_label']]
                cmp_rows.append(excel_row)

        # ---------- Sheet 2: 每日资产 ----------
        assets = self._daily_assets
        dates_sorted = sorted(assets.keys())
        # 从 snapshots 反推每日 cash（取当日最后一个快照）
        if self.account and self.account.snapshots:
            daily_last_snap = {}
            for s in self.account.snapshots:
                d = s.created_at.date()
                daily_last_snap[d] = s
            asset_rows = []
            for d in dates_sorted:
                nav = assets[d]
                snap = daily_last_snap.get(d)
                cash = snap.cash if snap else 0
                pos_val = nav - cash
                asset_rows.append({
                    '日期': d.strftime('%Y-%m-%d'),
                    '现金': round(cash, 2),
                    '持仓市值': round(pos_val, 2),
                    '总净值': round(nav, 2),
                })
        else:
            asset_rows = [{'日期': d.strftime('%Y-%m-%d'), '总净值': round(v, 2)}
                         for d, v in sorted(assets.items())]

        # ---------- Sheet 3: 交易记录 ----------
        trade_rows = []
        has_notes = self.account and any(getattr(t, 'note', '') for t in self.account.trade_records)
        if self.account:
            for t in self.account.trade_records:
                row = {
                    '日期': t.created_at.strftime('%Y-%m-%d %H:%M'),
                    '标的': t.symbol,
                    '方向': '买入' if t.side == 1 else '卖出',
                    '价格': t.price,
                    '数量': t.volume,
                    '金额': round(t.amount, 2),
                    '手续费': round(t.fee, 2),
                }
                if has_notes:
                    row['备注'] = getattr(t, 'note', '')
                trade_rows.append(row)

        # ---------- 写入 Excel ----------
        wb = openpyxl.Workbook()

        # 样式定义
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=11, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def write_sheet(ws, title, headers, rows, start_row=1):
            """写入一个 sheet 分组，带标题和表头，从 start_row 开始"""
            # 标题行
            end_col = len(headers)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
            ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=13)
            # 表头
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row + 1, column=ci, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            # 数据行
            for ri, row in enumerate(rows, start_row + 2):
                for ci, h in enumerate(headers, 1):
                    val = row.get(h, '')
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = thin_border
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
            # 列宽自适应
            for ci, h in enumerate(headers, 1):
                max_len = max(len(str(h)), max((len(str(r.get(h, ''))) for r in rows), default=0)) + 2
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len, 30)
            return start_row + 2 + len(rows) + 1  # 返回下一组的起始行（留空行）

        # Sheet 1: 指标分析（基础信息 / 基准对比 / 收益 / 风险 / 交易）
        ws1 = wb.active
        ws1.title = '指标分析'
        m_headers = ['指标名称', '数值', '说明']
        row = 1
        # 基础信息
        row = write_sheet(ws1, '基础信息', m_headers, base_rows, start_row=row)
        # 基准对比
        if has_bench:
            c_headers = ['指标名称', '策略', bench_label]
            row = write_sheet(ws1, f'策略 vs {bench_label}', c_headers, cmp_rows, start_row=row)
        # 收益 / 风险 / 交易
        for group_title, rows in [('收益指标', groups.get('收益', [])),
                                  ('风险指标', groups.get('风险', [])),
                                  ('交易指标', groups.get('交易', []))]:
            if not rows:
                continue
            if group_title == '交易指标' and not has_records:
                continue
            row = write_sheet(ws1, group_title, m_headers, rows, start_row=row)

        # Sheet 2: 每日资产
        ws2 = wb.create_sheet('每日资产')
        a_headers = list(asset_rows[0].keys()) if asset_rows else []
        write_sheet(ws2, f'{report_name} — 每日资产', a_headers, asset_rows)

        # Sheet 3: 交易记录
        ws3 = wb.create_sheet('交易记录')
        t_headers = list(trade_rows[0].keys()) if trade_rows else []
        write_sheet(ws3, f'{report_name} — 交易记录', t_headers, trade_rows)

        # 保存
        from pathlib import Path
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(self.base_dir) / output_dir / f"{report_name}_{current_datetime}.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f'Excel 已生成至: {output_path}')
        return str(output_path)

    @staticmethod
    def _aggregate_daily_assets(snapshots: List) -> Dict:
        """
        聚合账户快照为日数据（每日取最后一个快照）
        
        账户类每次交易都会生成快照，一天内可能有多次交易
        此方法将日内多个快照聚合为每日一个数据点，取最后一个快照的资产值
        
        Args:
            snapshots: AccountSnapshot 对象列表
                      每个快照包含 created_at（datetime）和 nav（资产净值）
            
        Returns:
            Dict[date, float]: {日期：资产净值} 字典
                              日期为 date 对象，资产净值为 float
            
        Example:
            >>> daily = AccountAnalyzer._aggregate_daily_assets(account.snapshots)
            >>> # 结果：{date(2024, 1, 1): 100000, date(2024, 1, 2): 101500, ...}
        """
        daily_snapshots = defaultdict(list)
        for snapshot in snapshots:
            snapshot_date = snapshot.created_at.date()
            daily_snapshots[snapshot_date].append(snapshot)

        result = {
            snapshot_date: snaps[-1].nav
            for snapshot_date, snaps in daily_snapshots.items()
        }

        return dict(sorted(result.items()))


    def _calculate_profit(self, trade_records: List) -> List:
        """
        计算每笔交易的盈亏
        
        使用 FIFO（先进先出）原则匹配开平仓交易：
        - 买入：累积持仓，记录平均开仓价格和手续费
        - 卖出：按持仓比例计算成本和盈亏，平仓后重置持仓信息
        
        盈亏计算公式：
        盈亏 = 卖出金额 - 成本 - 手续费
        成本 = (卖出数量 / 持仓数量) × 总成本
        
        Args:
            trade_records: 成交记录列表
                         每条记录包含：symbol, volume, price, side, fee, created_at
            
        Returns:
            List[Dict]: 盈亏记录列表，每条记录包含：
                       - symbol: 标的代码
                       - profit: 盈亏金额（正为盈利，负为亏损）
                       - open_time, close_time: 开平仓时间
                       - open_price, close_price: 开平仓价格
                       - volume: 交易数量（负数表示卖出）
                       - open_fee, close_fee: 开平仓手续费
                       - original_trade: 原始交易记录对象
            
        Example:
            >>> profits = analyzer._calculate_profit(trade_records)
            >>> for trade in profits:
            ...     print(f"{trade['symbol']}: 盈亏 {trade['profit']}")
        """
        positions = defaultdict(lambda: {'volume': 0, 'cost': 0, 'open_time': None, 'open_price': 0, 'open_fee': 0})
        processed_trades = []

        for trade in trade_records:
            if trade.volume == 0 or math.isnan(trade.price):
                continue
            symbol = trade.symbol
            volume = trade.volume
            abs_volume = abs(volume)
            price = trade.price
            side = trade.side
            created_at = trade.created_at
            fee = trade.fee

            # [修复] 2026-05-30 TradeRecord.side 是 int (OrderSide.Buy=1)，
            #   原用字符串 'buy'/'sell' 比较导致永久为 False，交易盈亏始终为空
            if side == OrderSide.Buy:
                positions[symbol]['volume'] += abs_volume
                positions[symbol]['cost'] += abs_volume * price + fee
                if positions[symbol]['open_time'] is None:
                    positions[symbol]['open_time'] = created_at
                    positions[symbol]['open_price'] = price
                    positions[symbol]['open_fee'] += fee
            elif side == OrderSide.Sell:
                if positions[symbol]['volume'] == 0:
                    continue

                sell_amount = abs_volume * price
                cost = (abs_volume / positions[symbol]['volume']) * positions[symbol]['cost']
                profit = sell_amount - cost - fee

                open_fee_portion = (abs_volume / positions[symbol]['volume']) * positions[symbol]['open_fee']

                processed_trades.append({
                    'symbol': symbol,
                    'profit': profit,
                    'open_time': positions[symbol]['open_time'],
                    'close_time': created_at,
                    'open_price': positions[symbol]['open_price'],
                    'open_fee': open_fee_portion,
                    'close_fee': fee,
                    'close_price': price,
                    'volume': volume,
                    'original_trade': trade
                })

                positions[symbol]['volume'] -= abs_volume
                positions[symbol]['cost'] -= cost
                positions[symbol]['open_fee'] -= open_fee_portion

                if positions[symbol]['volume'] == 0:
                    positions[symbol]['open_time'] = None
                    positions[symbol]['open_price'] = 0
                    positions[symbol]['open_fee'] = 0

        return processed_trades

    def _get_caller_dir(self) -> str:
        """
        获取调用者所在目录
        
        使用 inspect 模块获取调用当前方法的代码所在的目录
        用于确定 HTML 报告的输出路径基准
        
        Returns:
            str: 调用者脚本的绝对目录路径
            
        Example:
            # 如果在 d:\\project\\backtest\\main.py 中调用
            # analyzer._get_caller_dir() 返回 "d:\\project\\backtest"
        """
        frame = inspect.currentframe()
        try:
            caller_frame = None
            for frame_info in inspect.stack():
                if frame_info.filename != __file__:
                    caller_frame = frame_info
                    break

            if caller_frame:
                return os.path.dirname(os.path.abspath(caller_frame.filename))
            else:
                return os.path.dirname(os.path.abspath(__file__))
        finally:
            del frame
